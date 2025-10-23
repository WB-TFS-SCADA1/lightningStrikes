import pandas as pd
import numpy as np
from geopy.distance import geodesic
from typing import List, Tuple, Dict
from datetime import datetime, timedelta
import pytz
from dotenv import dotenv_values
import csv
import pathlib
import os
import sys
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
import smtplib, ssl
from email import encoders
import logging
from logging.config import dictConfig
import traceback
import requests  # Add this import at the top with other imports
from openpyxl import Workbook  # Added for Excel
from openpyxl.styles import Alignment # Added for cell alignment/word wrap
from openpyxl.utils import get_column_letter # Added for column width adjustment

if getattr(sys, 'frozen', False):
    currentDir = os.path.dirname(sys.executable)
    currentFile = os.path.basename(sys.executable)
elif __file__:
    currentDir = pathlib.Path(__file__).parent.resolve()
    currentFile = os.path.basename(__file__)
else:
    currentDir = pathlib.Path(__file__).parent.resolve()
    currentFile = os.path.basename(__file__)

config = dotenv_values(f"{currentDir}/.env")



def getLogConf(logName):
    dictConf = {
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'standard': {
                'format': "[%(asctime)s] %(levelname)s [%(name)s.%(funcName)s:%(lineno)d] %(message)s",
                'datefmt': '%Y-%m-%dT%H:%M:%S',
            },
        },
        'handlers': {
            'default': {
                'level': 'DEBUG',
                'class': 'logging.StreamHandler',
                'formatter': 'standard',
                'stream': sys.stderr,
            },
            'rotating_to_file': {
                'level': 'DEBUG',
                'class': "logging.handlers.RotatingFileHandler",
                'formatter': 'standard',
                "filename": logName,
                "maxBytes": 10000000,
                "backupCount": 2,
            },
        },
        'loggers': {
            '': {
                'handlers': ['default', 'rotating_to_file'],
                'level': 'DEBUG',
                'propagate': True
            }
        }
    }
    return dictConf


logName = f'{currentDir}/logs/{currentFile.replace(".py", "")}.log'
loggingConfig = getLogConf(logName)
logging.config.dictConfig(loggingConfig)
logger = logging.getLogger(__name__)

logger.info(currentDir)


def emailReport(filename, emailMsg, emailSubject):
    logger.info(f"Attempting to send email report: {emailSubject}")

    try:
        # Create message
        msg = MIMEMultipart()
        msg['Subject'] = emailSubject
        msg['From'] = config['emailUser']
        msg.attach(MIMEText(emailMsg, 'html'))

        # Attach file
        logger.info(f"Attaching file: {filename}")
        with open(filename, "rb") as attachment:
            if filename.endswith(".xlsx"):
                part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(filename)}",
                )
            else: # Assuming CSV or other, fallback to octet-stream
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(filename)}",
                )
        msg.attach(part)

        context = ssl.create_default_context()

        logger.info(f"Connecting to SMTP server: {config['smtpServer']}:{config['smtpPort']}")

        # Connect to server with timeout
        with smtplib.SMTP(config['smtpServer'], int(config['smtpPort']), timeout=120) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()

            logger.info("Logging into SMTP server")
            server.login(config['emailUser'], config['emailPass'])

            # Send the email
            logger.info("Sending email")
            server.sendmail(
                config['emailUser'],
                config['businessEmails'].split(','),
                msg.as_string()
            )
            logger.info("Email sent successfully")

    except ssl.SSLError as e:
        logger.error(f"SSL Error occurred: {str(e)}")
        raise
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error occurred: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error sending email: {str(e)}")
        logger.error(traceback.format_exc())
        raise


def validate_coordinates(lat: float, lon: float) -> bool:
    """
    Validate that coordinates are within valid ranges
    """
    try:
        lat_float = float(lat)
        lon_float = float(lon)
        return -90 <= lat_float <= 90 and -180 <= lon_float <= 180
    except (ValueError, TypeError):
        return False

def load_data(site_query: str, strikes_query: str, query_params: Tuple) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load data from SQL Server database (sites) and PostgreSQL database (lightning strikes)
    """
    import pyodbc
    import psycopg2
    
    # Connect to SQL Server for sites
    sql_server_conn = pyodbc.connect(
        f'DRIVER={{ODBC Driver 17 for SQL Server}};'
        f'SERVER={config["sqlServer"]};'
        f'DATABASE={config["sqlDatabase"]};'
        f'UID={config["sqlUser"]};'
        f'PWD={config["sqlPassword"]}'
    )
    
    # Connect to PostgreSQL for lightning strikes
    pg_conn = psycopg2.connect(
        host=config['PG_HOST'],
        database=config['PG_DATABASE'],
        user=config['PG_USER'],
        password=config['PG_PASSWORD']
    )
    
    # Load the data
    sites_df = pd.read_sql(site_query, sql_server_conn)
    strikes_df = pd.read_sql(strikes_query, pg_conn, params=query_params)
    
    # Convert UTC timestamps to Central Time
    central = pytz.timezone('America/Chicago')
    strikes_df['Timestamp'] = pd.to_datetime(strikes_df['Timestamp'])
    strikes_df['Timestamp'] = strikes_df['Timestamp'].apply(
        lambda x: x.replace(tzinfo=pytz.UTC).astimezone(central)
    )
    
    sql_server_conn.close()
    pg_conn.close()
    return sites_df, strikes_df


def get_strikes_for_site(site_row: pd.Series, strikes_df: pd.DataFrame, radius_miles: float) -> List[Dict]:
    """
    Get all strikes within radius of a site with their details
    """
    try:
        site_lat = float(site_row['Latitude'])
        site_lon = float(site_row['Longitude'])

        if not validate_coordinates(site_lat, site_lon):
            print(f"Warning: Invalid coordinates for site {site_row['SiteName']}: {site_lat}, {site_lon}")
            return []

        site_coords = (site_lat, site_lon)  # Note: order is (lat, lon)
    except (ValueError, TypeError) as e:
        print(f"Error parsing coordinates for site {site_row['SiteName']}: {e}")
        return []

    strikes_in_radius = []

    for _, strike in strikes_df.iterrows():
        try:
            strike_lat = float(strike['Latitude'])
            strike_lon = float(strike['Longitude'])

            if not validate_coordinates(strike_lat, strike_lon):
                continue

            strike_coords = (strike_lat, strike_lon)  # Note: order is (lat, lon)
            distance = geodesic(site_coords, strike_coords).miles

            if distance <= radius_miles:
                strikes_in_radius.append({
                    'latitude': strike_lat,
                    'longitude': strike_lon,
                    'timestamp': strike['Timestamp'],
                    'distance': distance,
                    'peakamp': strike['PeakAmp']  # Added PeakAmp to the dictionary
                })
        except (ValueError, TypeError) as e:
            print(f"Error processing strike coordinates: {e}")
            continue

    # Sort strikes by timestamp
    return sorted(strikes_in_radius, key=lambda x: x['timestamp'])


def create_detailed_report(sites_df: pd.DataFrame, strikes_df: pd.DataFrame, radii: List[float], filename: str) -> int:
    """
    Create a detailed report with strikes listed under each site
    Returns the number of sites with strikes
    """
    sites_with_strikes = 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Lightning Report"

    # Write header row
    header = ['Site Name', 'Latitude', 'Longitude',
              f'Strikes ({radii[0]} mi)']
    ws.append(header)

    for _, site in sites_df.iterrows():
        # Get strikes for each radius
        strikes_1mi = get_strikes_for_site(site, strikes_df, radii[0])

        # Skip sites with no strikes in either radius
        if not strikes_1mi:
            continue

        sites_with_strikes += 1

        # Write site summary row
        ws.append([
            site['SiteName'],
            site['Latitude'],
            site['Longitude'],
            len(strikes_1mi)
        ])

        # Write 1-mile radius strikes
        if strikes_1mi:
            ws.append(['Strikes within 1 mile:'])
            for strike in strikes_1mi:
                ws.append([
                    '  Strike',
                    strike['latitude'],
                    strike['longitude'],
                    strike['timestamp'].strftime('%Y-%m-%d %I:%M:%S %p %Z'),
                    f"{strike['distance']:.2f} miles",
                    f"{strike['peakamp']} kA"
                ])

        # Add blank line between sites
        ws.append([])

    # Auto-size columns for better readability
    for col_idx, column_cells in enumerate(ws.columns):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value is not None:
                    lines = str(cell.value).split('\n')
                    cell_len = 0
                    if lines:
                        cell_len = max(len(line) for line in lines)
                    
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass 
        
        column_letter = get_column_letter(col_idx + 1)
        adjusted_width = (max_length + 2) 
        if adjusted_width > 50: 
             adjusted_width = 50
        if adjusted_width < 10 and column_letter == 'A':
            adjusted_width = 10
        elif adjusted_width < 5:
             adjusted_width = 5

        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    return sites_with_strikes

def get_work_orders():
    """
    Fetch work orders from the API
    """
    logger.info("Fetching work orders from API")
    try:
        response = requests.get("https://wbrapi.azurewebsites.net/api/Fiix/WorkOrder")
        response.raise_for_status()
        return response.json()
    except Exception as e:
        logger.error(f"Error fetching work orders: {str(e)}")
        raise

def create_correlation_report(sites_df: pd.DataFrame, strikes_df: pd.DataFrame, work_orders: list, 
                            radii: List[float], filename: str) -> int:
    """
    Create a report correlating lightning strikes with work orders
    Only includes strikes that have associated work orders
    """
    sites_with_data = 0
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lightning WO Correlation"

    # Convert work orders to DataFrame and parse dates with UTC timezone
    wo_df = pd.DataFrame(work_orders)
    wo_df['createdDateTime'] = pd.to_datetime(wo_df['createdDateTime'], format='ISO8601', utc=True)
    
    # Filter work orders to last 14 days (using UTC)
    cutoff_date = pd.Timestamp.now(tz='UTC') - pd.Timedelta(days=14)
    wo_df = wo_df[wo_df['createdDateTime'] >= cutoff_date]

    # Write header row
    header = ['Site Name', 'Latitude', 'Longitude', f'Strikes ({radii[0]} mi)']
    ws.append(header)

    # Define column index for Work Order Description (0-indexed)
    # Assuming header for work orders is ['Work Order Number', 'Asset Name', 'Maintenance Type', 'Work Order Description', 'Created Date']
    # The full row for work orders is ['', 'Work Order Number', 'Asset Name', 'Maintenance Type', 'Work Order Description', 'Created Date']
    # So, if 'Work Order Description' is the 4th item in its own header, and there's an empty cell before the WO details, it will be column E (index 4)
    work_order_desc_col_letter = 'E' # This might need adjustment based on actual output structure

    for _, site in sites_df.iterrows():
        all_strikes = get_strikes_for_site(site, strikes_df, radii[0])
        if not all_strikes:
            continue
        
        site_work_orders = wo_df[wo_df['facilityID'] == site['facilityid']]
        if site_work_orders.empty:
            continue
            
        strikes_with_orders = []
        for strike in all_strikes:
            strike_time = strike['timestamp']
            if not strike_time.tzinfo:
                strike_time = pd.Timestamp(strike_time, tz='UTC')
            else:
                strike_time = pd.Timestamp(strike_time).tz_convert('UTC')
            
            matching_orders = site_work_orders[
                site_work_orders['createdDateTime'] >= strike_time
            ]
            if not matching_orders.empty:
                strikes_with_orders.append(strike)
        
        if not strikes_with_orders:
            continue
                
        sites_with_data += 1
        
        ws.append([
            site['SiteName'],
            site['Latitude'],
            site['Longitude'],
            len(strikes_with_orders)
        ])

        ws.append(['Strikes within 1 mile:'])
        for strike in strikes_with_orders:
            ws.append([
                '  Strike',
                strike['latitude'],
                strike['longitude'],
                strike['timestamp'].strftime('%Y-%m-%d %I:%M:%S %p %Z'),
                f"{strike['distance']:.2f} miles",
                f"{strike['peakamp']} kA"
            ])
        
        ws.append(['Work Orders Past 14 Days:'])
        wo_header = ['  Work Order Number', 'Asset Name', 'Maintenance Type', 
                       'Work Order Description', 'Created Date']
        ws.append(wo_header)
        
        earliest_strike = min(strike['timestamp'] for strike in strikes_with_orders)
        if not earliest_strike.tzinfo:
            earliest_strike = pd.Timestamp(earliest_strike, tz='UTC')
        else:
            earliest_strike = pd.Timestamp(earliest_strike).tz_convert('UTC')
        
        relevant_orders = site_work_orders[
            site_work_orders['createdDateTime'] >= earliest_strike
        ]
        
        for _, wo in relevant_orders.iterrows():
            # Filter out work orders containing the specified phrase
            if "Could not send Truck Unloading tags to " in str(wo['workOrderDesc']):
                continue

            created_time = wo['createdDateTime'].tz_convert('America/Chicago')
            # work_order_desc_excel_friendly = str(wo['workOrderDesc']).replace('\n', ' ').replace('\r', '') # No longer needed, handled by word wrap
            row_data = [
                f"  {wo['woNumber']}",
                wo['assetName'],
                wo['maintenanceType'],
                str(wo['workOrderDesc']), # Keep original newlines for word wrap
                created_time.strftime('%Y-%m-%d %I:%M:%S %p %Z')
            ]
            ws.append(row_data)
            # Apply word wrap to the Work Order Description cell
            # The cell is in the current last row, and the column letter is determined above.
            # Need to find the column index for 'Work Order Description' in wo_header
            try:
                desc_col_index = wo_header.index('Work Order Description')
                # Convert 0-indexed column to 1-indexed for openpyxl cell access
                # The first element in row_data (woNumber) might be indented, so direct indexing might be off.
                # Let's find the actual column letter for 'Work Order Description' based on the `row_data` structure.
                # If wo_header is [A, B, C, D, E] and row_data is [A, B, C, D, E]
                # And 'Work Order Description' is D in wo_header, then it's the 4th item in row_data.
                # openpyxl columns are 1-indexed. So, column D is 4.
                # The items in wo_header are: '  Work Order Number', 'Asset Name', 'Maintenance Type', 'Work Order Description', 'Created Date'
                # The items in row_data are: f"  {wo['woNumber']}", wo['assetName'], wo['maintenanceType'], str(wo['workOrderDesc']), created_time.strftime(...)
                # So 'Work Order Description' is the 4th element (index 3) in row_data.
                # Therefore, the column is 3 + 1 = 4 (D)
                target_cell = ws.cell(row=ws.max_row, column=desc_col_index + 1) # desc_col_index is 0-based from wo_header
                target_cell.alignment = Alignment(wrap_text=True, vertical='top')
            except ValueError:
                logger.warning("Could not find 'Work Order Description' in work order header for wrapping.")

        ws.append([])
    
    # Auto-size columns for better readability
    for col_idx, column_cells in enumerate(ws.columns):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value is not None:
                    lines = str(cell.value).split('\n')
                    cell_len = 0
                    if lines:
                        cell_len = max(len(line) for line in lines)
                    
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass 
        
        column_letter = get_column_letter(col_idx + 1)
        adjusted_width = (max_length + 2) 
        if adjusted_width > 50: 
             adjusted_width = 50
        if adjusted_width < 10 and column_letter == 'A':
            adjusted_width = 10
        elif adjusted_width < 5:
             adjusted_width = 5

        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    return sites_with_data

def main():
    try:
        # SQL queries for 7-day report
        site_query = """
        SELECT SiteName, Latitude, Longitude, site.id, facilityid
        FROM site 
        WHERE Latitude IS NOT NULL AND Longitude IS NOT NULL and site.type in ('Salt Water Disposal', 'Truck Transfer') and site.enabled = 1
        """

        strikes_query_7d = """
        SELECT latitude AS Latitude, longitude AS Longitude, peak_amp AS PeakAmp, strike_timestamp AS Timestamp
        FROM public.lightning_strikes
        WHERE latitude IS NOT NULL 
        AND longitude IS NOT NULL 
        AND pulse_type = 'cg'
        AND strike_timestamp >= NOW() - INTERVAL '7 days'
        """

        # Load data for 7-day report
        sites_df, strikes_df_7d = load_data(site_query, strikes_query_7d, ())

        # Generate filename with timestamp
        timestamp = datetime.now(pytz.timezone('America/Chicago')).strftime('%Y%m%d_%H%M%S')
        filename = (f'{currentDir}/detailed_lightning_report_{timestamp}.xlsx')

        # Create the 7-day report
        sites_with_strikes = create_detailed_report(sites_df, strikes_df_7d, [1.0], filename)

        # SQL query for 14-day strikes
        strikes_query_14d = """
        SELECT latitude AS Latitude, longitude AS Longitude, peak_amp AS PeakAmp, strike_timestamp AS Timestamp
        FROM public.lightning_strikes
        WHERE latitude IS NOT NULL 
        AND longitude IS NOT NULL 
        AND strike_timestamp >= NOW() - INTERVAL '14 days'
        """

        # Load data for 14-day correlation report
        _, strikes_df_14d = load_data(site_query, strikes_query_14d, ())

        # Get work orders
        work_orders = get_work_orders()
        print("=============================")
        print(len(work_orders))
        print("=============================")
        
        # Generate correlation report filename
        correlation_filename = f'{currentDir}/lightning_strike_wo_correlation_report_{timestamp}.xlsx'
        
        # Create correlation report with 14-day strike data
        sites_with_data = create_correlation_report(
            sites_df, 
            strikes_df_14d,  # Using 14-day strike data
            work_orders, 
            [1.0], 
            correlation_filename
        )

        # Print summary
        central_now = datetime.now(pytz.timezone('America/Chicago'))
        utc_now = central_now.astimezone(pytz.UTC)

        logger.info(f"\nReports saved as:")
        logger.info(f"7-day report: {filename}")
        logger.info(f"14-day correlation report: {correlation_filename}")
        logger.info(f"\n7-day report summary:")
        logger.info(f"Total lightning strikes analyzed: {len(strikes_df_7d)}")
        logger.info(f"Total sites analyzed: {len(sites_df)}")
        logger.info(f"Sites with strikes: {sites_with_strikes}")
        logger.info(f"Sites without strikes: {len(sites_df) - sites_with_strikes}")
        logger.info(f"\n14-day correlation report summary:")
        logger.info(f"Total lightning strikes analyzed: {len(strikes_df_14d)}")
        logger.info(f"Sites with correlated data: {sites_with_data}")
        logger.info(f"\nDate ranges:")
        logger.info(f"7-day report: Last 7 days from {central_now.strftime('%Y-%m-%d %I:%M:%S %p %Z')}")
        logger.info(f"14-day report: Last 14 days from {central_now.strftime('%Y-%m-%d %I:%M:%S %p %Z')}")
        
        # Send both reports
        emailReport(filename,
                   f"Lightning report (7-day) attached.",
                   f"Weekly Lightning Strike Report")
        
        emailReport(correlation_filename,
                   f"Lightning strike and work order correlation report (14-day) attached.",
                   f"Weekly Lightning Strike and Work Order Correlation Report")
                   
    except Exception as e:
        print(f"Script {currentFile} has failed:\n {traceback.format_exc()}", 'html')

        logger.info(traceback.format_exc())
        msg = MIMEMultipart()
        msg['Subject'] = f"Script {currentFile} - Failed"
        msg['From'] = config['emailUser']
        context = ssl.create_default_context()
        server = smtplib.SMTP(config['smtpServer'], int(config['smtpPort']), timeout=120)
        server.ehlo()
        server.starttls(context=context)
        server.ehlo()
        server.login(config['emailUser'], config['emailPass'])
        try:
            logger.info(f"Program failed")
            logger.info(f"error: {traceback.format_exc()}")

            msgErr = MIMEMultipart()
            msgErr['Subject'] = f"Script {currentFile} - Failed"
            msgErr['To'] = config['errorEmails']
            msgErr['From'] = config['emailUser']
            msgErr.attach(MIMEText(f"Script {currentFile} has failed:\n {traceback.format_exc()}", 'html'))
            server.sendmail(config['emailUser'], config['errorEmails'], msgErr.as_string())
            server.quit()
        except:
            msgErr = MIMEMultipart()
            msgErr['Subject'] = f"Script {currentFile} - Failed"
            msgErr['To'] = config['errorEmails']
            msgErr['From'] = config['emailUser']
            msgErr.attach(MIMEText(f"Script {currentFile} has failed:\n {traceback.format_exc()}", 'html'))
            server.sendmail(config['emailUser'], config['errorEmails'], msgErr.as_string())
            server.quit()

if __name__ == "__main__":
    main()